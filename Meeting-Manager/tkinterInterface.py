# import openpyxl and tkinter modules
from tkinter import filedialog

from openpyxl import load_workbook
from tkinter import *
from Main import my_logic

# globally declare wb and sheet variable

# opening the existing excel file

wb = load_workbook('data.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Meeting Title"
    sheet.cell(row=1, column=2).value = "Meeting Transcript"
    sheet.cell(row=1, column=3).value = "Enterprise ID"
    sheet.cell(row=1, column=4).value = "Email ID"
    sheet.cell(row=1, column=5).value = "Meeting Minutes"
    sheet.cell(row=1, column=6).value = "Participant List"
    sheet.cell(row=1, column=7).value = "Follow-Up Meeting"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the transcript_field box
    transcript_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the eid_field box
    eid_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the email_field box
    email_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the minutes_field box
    minutes_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the participants_field box
    participants_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the followup_field box
    followup_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    title_field.delete(0, END)
    # transcript_field.delete(0, END)
    eid_field.delete(0, END)
    email_field.delete(0, END)
    # minutes_field.delete(0, END)
    # participants_field.delete(0, END)
    minutes_field.deselect()
    participants_field.deselect()
    followup_field.deselect()


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (title_field.get() == "" and
            textfile.get() == "" and
            eid_field.get() == "" and
            email_field.get() == ""):
        # minutes_field.get() == "" and
        # participants_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = title_field.get()
        sheet.cell(row=current_row + 1, column=2).value = textfile
        sheet.cell(row=current_row + 1, column=3).value = eid_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = Checkbutton1.get()
        sheet.cell(row=current_row + 1, column=6).value = Checkbutton2.get()
        sheet.cell(row=current_row + 1, column=7).value = Checkbutton3.get()

        my_logic(textfile, Checkbutton1.get(), Checkbutton2.get(), Checkbutton3.get())

        # save the file
        wb.save('data.xlsx')

        # set focus on the name_field box
        title_field.focus_set()

        # call the clear() function
        clear()


# File Explorer to Upload File
def openFile():
    global textfile
    filename = filedialog.askopenfilename(initialdir="/", title="Upload File",
                                          filetypes=(("Text files", "*.txt"), ("all files", "*.*")))
    upload_label.configure(text="Uploaded: " + filename)
    file = open(filename, "r")
    tf = file.read()
    textfile = tf



    file.close()


# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='lavender')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("1000x1000")

    excel()

    # create a Form label
    heading = Label(root, text="Meeting Manager", bg="lavender")

    # create a Meeting Title label
    title = Label(root, text="Meeting Title", bg="lavender")

    # create a Meeting Transcript label
    transcript = Label(root, text="Meeting Transcript", bg="lavender")

    # create a Enterprise ID label
    eid = Label(root, text="Enterprise ID", bg="lavender")

    # create a Email ID label
    email = Label(root, text="Email ID", bg="lavender")

    # create a Meeting Minutes label
    minutes = Label(root, text=" ", bg="lavender")

    # create a Participant List label
    participants = Label(root, text=" ", bg="lavender")

    # create a Follow-Up Meeting Label
    followup = Label(root, text=" ", bg="lavender")

    # Upload Label
    upload_label = Label(root, text="Upload Pending", bg="lavender")

    # saveButton = Button(root, text="Save Choices", width=20, height=2, command=savecheck)
    Checkbutton1 = IntVar()
    Checkbutton2 = IntVar()
    Checkbutton3 = IntVar()
    textfile = StringVar()

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    title.grid(row=1, column=0)
    transcript.grid(row=2, column=0)
    eid.grid(row=3, column=0)
    email.grid(row=4, column=0)
    minutes.grid(row=5, column=0)
    # saveButton.grid(row=7, column=0)
    participants.grid(row=6, column=0)
    followup.grid(row=7, column=0)
    # saveButton.grid(row=7, column=1)
    upload_label.grid(row=14, column=1)

    # create a text entry box
    # for typing the information
    title_field = Entry(root)
    transcript_field = Button(root, text="Upload Transcript", command=openFile)
    eid_field = Entry(root)
    email_field = Entry(root)
    minutes_field = Checkbutton(root, text="Generate Meeting Minutes", variable=Checkbutton1, onvalue=1, offvalue=0,
                                height=2, width=10)
    participants_field = Checkbutton(root, text="Generate Participant List", variable=Checkbutton2, onvalue=1,
                                     offvalue=0, height=2, width=10)
    followup_field = Checkbutton(root, text="Schedule Follow-Up Meeting", variable=Checkbutton3, onvalue=1,
                                 offvalue=0, height=2, width=10)
    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    title_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    transcript_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus5 function
    eid_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus3 function
    minutes_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus4 function
    participants_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus5 function
    followup_field.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure.
    title_field.grid(row=1, column=1, ipadx="100")
    transcript_field.grid(row=2, column=1, ipadx="100")
    eid_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    minutes_field.grid(row=5, column=1, ipadx="100")
    participants_field.grid(row=6, column=1, ipadx="100")
    followup_field.grid(row=7, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="light green", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()
